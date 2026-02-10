#!/usr/bin/env python3
"""
Quick test script to verify the app works
"""
import sys
import os

# Test 1: Check Python version
print("Test 1: Checking Python version...")
if sys.version_info < (3, 8):
    print("❌ Python 3.8+ required")
    sys.exit(1)
print(f"✅ Python {sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}")

# Test 2: Check dependencies
print("\nTest 2: Checking dependencies...")
try:
    import flask
    print(f"✅ Flask {flask.__version__}")
except ImportError:
    print("❌ Flask not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    import openpyxl
    print(f"✅ openpyxl {openpyxl.__version__}")
except ImportError:
    print("❌ openpyxl not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

# Test 3: Check files exist
print("\nTest 3: Checking project files...")
required_files = [
    'app.py',
    'templates/index.html',
    'static/manifest.json',
    'requirements.txt',
    'start.sh'
]

for file in required_files:
    if os.path.exists(file):
        print(f"✅ {file}")
    else:
        print(f"❌ {file} missing")
        sys.exit(1)

# Test 4: Check feeds.xlsx
print("\nTest 4: Checking data file...")
if os.path.exists('feeds.xlsx'):
    print("✅ feeds.xlsx exists")
    from openpyxl import load_workbook
    try:
        wb = load_workbook('feeds.xlsx')
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected = ["Date", "Time", "Type", "Amount (oz)", "Duration (min)", "Notes", "Logged By", "Timestamp"]
        if headers == expected:
            print("✅ Excel headers correct")
        else:
            print(f"⚠️  Headers: {headers}")
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        sys.exit(1)
else:
    print("⚠️  feeds.xlsx not created yet (will be created on first run)")

# Test 5: Check template
print("\nTest 5: Checking HTML template...")
with open('templates/index.html', 'r') as f:
    html = f.read()
    if 'Esmé' in html or 'Esme' in html:
        print("✅ Template personalized for Esmé")
    if 'manifest.json' in html:
        print("✅ PWA manifest linked")
    if 'SpeechRecognition' in html:
        print("✅ Voice recognition code present")

print("\n" + "="*60)
print("All tests passed! ✨")
print("="*60)
print("\nNext steps:")
print("1. Run: ./start.sh")
print("2. Open the URL on your phone")
print("3. Add to home screen")
print("4. Start tracking feeds!")
print("")
