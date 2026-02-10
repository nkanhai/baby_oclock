"""
Test concurrent write safety - critical for two parents logging simultaneously.
"""

import pytest
import threading
from openpyxl import load_workbook


class TestConcurrentWrites:
    """Test simultaneous write operations"""

    def test_two_simultaneous_posts(self, client, temp_xlsx):
        """Two simultaneous POSTs both succeed"""
        results = []

        def post_feed(amount):
            try:
                resp = client.post('/api/feeds', json={
                    "type": "bottle",
                    "amount_ml": amount,
                    "logged_by": "Mom"
                })
                results.append(('success', resp.status_code, amount))
            except Exception as e:
                results.append(('error', str(e), amount))

        # Fire two POSTs simultaneously
        threads = [
            threading.Thread(target=post_feed, args=(90.0,)),
            threading.Thread(target=post_feed, args=(75.0,))
        ]

        for t in threads:
            t.start()
        for t in threads:
            t.join()

        # Both should succeed
        assert len(results) == 2
        assert all(r[0] == 'success' and r[1] == 201 for r in results)

        # Verify both entries exist in xlsx
        wb = load_workbook(temp_xlsx)
        ws = wb.active

        # Should have 3 rows: 1 header + 2 data
        assert ws.max_row == 3

        # Verify both amounts are present
        amounts = [row[3] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert 90.0 in amounts
        assert 75.0 in amounts

    def test_five_rapid_fire_posts(self, client, temp_xlsx):
        """Five threads POST at once"""
        results = []

        def post_feed(amount):
            resp = client.post('/api/feeds', json={
                "type": "bottle",
                "amount_ml": amount,
                "logged_by": "Mom" if amount % 2 == 0 else "Dad"
            })
            results.append(resp.status_code)

        # Fire 5 POSTs simultaneously
        # Using ml values 30, 60, 90, 120, 150
        threads = [threading.Thread(target=post_feed, args=(i * 30,)) for i in range(1, 6)]

        for t in threads:
            t.start()
        for t in threads:
            t.join()

        # All should succeed
        assert all(code == 201 for code in results)

        # Verify all entries exist in xlsx
        wb = load_workbook(temp_xlsx)
        ws = wb.active

        # Should have 6 rows: 1 header + 5 data
        assert ws.max_row == 6

    def test_read_during_write(self, client, temp_xlsx):
        """GET during POST returns valid response"""
        results = {'post': None, 'get': None}

        def post_feeds():
            for i in range(10):
                resp = client.post('/api/feeds', json={
                    "type": "bottle",
                    "amount_ml": 60.0
                })
                results['post'] = resp.status_code

        def get_feeds():
            for i in range(10):
                resp = client.get('/api/feeds')
                if resp.status_code == 200:
                    results['get'] = 200

        # Run POST and GET in parallel
        t1 = threading.Thread(target=post_feeds)
        t2 = threading.Thread(target=get_feeds)

        t1.start()
        t2.start()
        t1.join()
        t2.join()

        # Both should succeed
        assert results['post'] == 201
        assert results['get'] == 200

    def test_ten_sequential_rapid_posts(self, client, temp_xlsx):
        """Ten rapid POSTs in sequence"""
        for i in range(10):
            resp = client.post('/api/feeds', json={
                "type": "bottle",
                "amount_ml": float((i + 1) * 30)
            })
            assert resp.status_code == 201

        # Verify all saved
        wb = load_workbook(temp_xlsx)
        ws = wb.active

        assert ws.max_row == 11  # 1 header + 10 data

    def test_no_duplicate_entries(self, client, temp_xlsx):
        """Single POST creates exactly one row"""
        # Count rows before
        client.post('/api/feeds', json={"type": "bottle", "amount_ml": 30.0})
        wb = load_workbook(temp_xlsx)
        rows_before = wb.active.max_row

        # Post once more
        client.post('/api/feeds', json={
            "type": "bottle",
            "amount_ml": 60.0
        })

        # Count rows after
        wb = load_workbook(temp_xlsx)
        rows_after = wb.active.max_row

        # Should be exactly 1 more row
        assert rows_after == rows_before + 1

    def test_concurrent_delete_and_create(self, client, temp_xlsx, seed_data):
        """Delete and create operations can happen concurrently"""
        # Get a feed to delete
        response = client.get('/api/feeds?date=2026-02-10')
        feed_id = response.get_json()['feeds'][0]['id']

        results = []

        def delete_feed():
            resp = client.delete(f'/api/feeds/{feed_id}')
            results.append(('delete', resp.status_code))

        def create_feed():
            resp = client.post('/api/feeds', json={
                "type": "bottle",
                "amount_ml": 150.0
            })
            results.append(('create', resp.status_code))

        # Run both operations
        t1 = threading.Thread(target=delete_feed)
        t2 = threading.Thread(target=create_feed)

        t1.start()
        t2.start()
        t1.join()
        t2.join()

        # Both should succeed
        assert len(results) == 2
        assert any(r[0] == 'delete' and r[1] == 200 for r in results)
        assert any(r[0] == 'create' and r[1] == 201 for r in results)
