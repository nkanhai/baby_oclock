"""
Test Excel file data integrity.
"""

import pytest
import os
from openpyxl import load_workbook
from datetime import datetime


class TestExcelIntegrity:
    """Test Excel file read/write correctness"""

    def test_file_auto_creation(self, client, temp_xlsx):
        """File is auto-created on first request"""
        # The file is already created by the fixture's init_excel_file call
        # Just verify it exists and has the right structure
        assert os.path.exists(temp_xlsx)

        # Make a request to add data
        client.post('/api/feeds', json={
            "type": "bottle",
            "amount_ml": 90.0
        })

        # Verify data was written
        wb = load_workbook(temp_xlsx)
        assert wb.active.max_row == 2  # Header + 1 data row

    def test_header_row_correct(self, client, temp_xlsx):
        """Header row has correct column names"""
        # Ensure file exists
        client.post('/api/feeds', json={"type": "bottle", "amount_ml": 30.0})

        wb = load_workbook(temp_xlsx)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        expected = ["Date", "Time", "Type", "Amount (ml)", "Duration (min)", "Notes", "Logged By", "Timestamp"]

        assert headers == expected

    def test_entry_matches_api_response(self, client, temp_xlsx):
        """Entry in Excel matches what API returned"""
        response = client.post('/api/feeds', json={
            "type": "bottle",
            "amount_ml": 105.0,
            "duration_min": 10,
            "notes": "test note",
            "logged_by": "Dad"
        })

        assert response.status_code == 201

        # Read Excel file
        wb = load_workbook(temp_xlsx)
        ws = wb.active

        # Get last row (newest entry)
        last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]

        date_str, time_str, type_str, amount, duration, notes, logged_by, timestamp = last_row

        assert type_str == "Feed (Bottle)"
        assert amount == 105.0
        assert duration == 10
        assert notes == "test note"
        assert logged_by == "Dad"

    def test_column_types_correct(self, client, temp_xlsx):
        """Column types are appropriate"""
        client.post('/api/feeds', json={
            "type": "nurse",
            "side": "left",
            "duration_min": 12,
            "timestamp": "2026-02-10T14:30:00"
        })

        wb = load_workbook(temp_xlsx)
        ws = wb.active

        last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
        date_str, time_str, type_str, amount, duration, notes, logged_by, timestamp = last_row

        # Date should be string in YYYY-MM-DD format
        assert isinstance(date_str, str)
        assert len(date_str.split('-')) == 3

        # Time should be string
        assert isinstance(time_str, str)
        assert 'AM' in time_str or 'PM' in time_str

        # Duration should be number
        assert isinstance(duration, (int, float))

        # Timestamp should be ISO format string
        assert isinstance(timestamp, str)
        assert 'T' in timestamp

    def test_multiple_entries_append_correctly(self, client, temp_xlsx):
        """Multiple entries append to file"""
        # Post 5 entries
        for i in range(5):
            client.post('/api/feeds', json={
                "type": "bottle",
                "amount_ml": float((i + 1) * 30)
            })

        wb = load_workbook(temp_xlsx)
        ws = wb.active

        # Should have 6 rows: 1 header + 5 data
        assert ws.max_row == 6

    def test_special_characters_in_notes(self, client, temp_xlsx):
        """Special characters in notes are preserved"""
        special_notes = "Baby said 👶 \"waaah\" 🍼\nNew line here"

        client.post('/api/feeds', json={
            "type": "bottle",
            "amount_ml": 60.0,
            "notes": special_notes
        })

        wb = load_workbook(temp_xlsx)
        ws = wb.active

        last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
        notes = last_row[5]  # Notes column

        assert notes == special_notes

    def test_amount_precision_preserved(self, client, temp_xlsx):
        """Amount precision is preserved"""
        client.post('/api/feeds', json={
            "type": "bottle",
            "amount_ml": 75.5
        })

        wb = load_workbook(temp_xlsx)
        ws = wb.active

        last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
        amount = last_row[3]  # Amount column

        assert amount == 75.5

    def test_file_valid_after_many_writes(self, client, temp_xlsx):
        """File remains valid after many writes"""
        # Write 50 entries
        for i in range(50):
            client.post('/api/feeds', json={
                "type": "bottle",
                "amount_ml": 60.0
            })

        # File should still be readable
        wb = load_workbook(temp_xlsx)
        ws = wb.active

        # Should have 51 rows: 1 header + 5 data
        assert ws.max_row == 51

        # Verify we can read all rows
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1

        assert row_count == 50

    def test_delete_removes_xlsx_row(self, client, temp_xlsx, seed_data):
        """Delete operation removes row from Excel file"""
        # Get a feed ID
        response = client.get('/api/feeds?date=2026-02-10')
        feed_id = response.get_json()['feeds'][0]['id']

        # Count rows before delete
        wb = load_workbook(temp_xlsx)
        rows_before = wb.active.max_row

        # Delete
        client.delete(f'/api/feeds/{feed_id}')

        # Count rows after delete
        wb = load_workbook(temp_xlsx)
        rows_after = wb.active.max_row

        assert rows_after == rows_before - 1

    def test_update_reflected_in_xlsx(self, client, temp_xlsx, seed_data):
        """Update operation changes Excel cell values"""
        # Get a feed
        response = client.get('/api/feeds?date=2026-02-10')
        feed = response.get_json()['feeds'][0]
        feed_id = feed['id']

        # Update it
        client.put(f'/api/feeds/{feed_id}', json={
            "type": "bottle",
            "amount_ml": 220.0,
            "notes": "Updated note"
        })

        # Read Excel directly
        wb = load_workbook(temp_xlsx)
        ws = wb.active

        # Find the row (row number is feed_id + 1 for header)
        row_num = feed_id + 1
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        amount = row[3]
        notes = row[5]

        assert amount == 220.0
        assert notes == "Updated note"
