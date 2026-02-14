#!/usr/bin/env python3
"""
Baby Feed Tracker - Flask Backend
Dead simple feed tracking for sleep-deprived parents.
"""

from flask import Flask, render_template, request, jsonify
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import threading
import os
import socket

app = Flask(__name__)


def parse_iso_timestamp(ts_string):
    """Parse ISO format timestamp, handling 'Z' suffix that Python 3.9 doesn't support."""
    if ts_string.endswith('Z'):
        ts_string = ts_string[:-1] + '+00:00'
    return datetime.fromisoformat(ts_string)

# Excel file path - configurable for testing
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FEED_FILE = os.path.join(BASE_DIR, 'feeds.xlsx')
app.config['FEED_FILE'] = os.environ.get('FEED_FILE', DEFAULT_FEED_FILE)

# Thread lock for file writes
file_lock = threading.Lock()


def get_excel_file():
    """Get the current Excel file path from app config."""
    return app.config.get('FEED_FILE', 'feeds.xlsx')


def get_local_ip():
    """Get the local network IP address."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "localhost"


def init_excel_file():
    """Create the Excel file with headers if it doesn't exist."""
    if not os.path.exists(get_excel_file()):
        with file_lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "Feed Log"

            # Headers
            headers = ["Date", "Time", "Type", "Amount (ml)", "Duration (min)", "Notes", "Logged By", "Timestamp"]
            ws.append(headers)

            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Set column widths
            ws.column_dimensions["A"].width = 12  # Date
            ws.column_dimensions["B"].width = 12  # Time
            ws.column_dimensions["C"].width = 18  # Type
            ws.column_dimensions["D"].width = 12  # Amount
            ws.column_dimensions["E"].width = 14  # Duration
            ws.column_dimensions["F"].width = 30  # Notes
            ws.column_dimensions["G"].width = 12  # Logged By
            ws.column_dimensions["H"].width = 20  # Timestamp

            wb.save(get_excel_file())
            print(f"✓ Created {get_excel_file()}")


def format_feed_type(feed_type, side=None):
    """Convert feed type and side into Excel-friendly string."""
    if feed_type == "bottle":
        if side == "formula":
            return "Feed (Bottle - Formula)"
        elif side == "milk":
            return "Feed (Bottle - Milk)"
        else:
            return "Feed (Bottle)"
    elif feed_type == "nurse":
        if side == "left":
            return "Nurse (Left)"
        elif side == "right":
            return "Nurse (Right)"
        else:
            return "Nurse (Both)"
    elif feed_type == "pump":
        if side == "left":
            return "Pump (Left)"
        elif side == "right":
            return "Pump (Right)"
        else:
            return "Pump (Both)"
    elif feed_type == "diaper":
        if side == "pee":
            return "Diaper (Pee)"
        elif side == "poop":
            return "Diaper (Poop)"
        elif side == "both":
            return "Diaper (Both)"
        else:
            return "Diaper"
    elif feed_type == "vitamin_d":
        return "Vitamin D"
    return feed_type


def add_feed_to_excel(feed_data):
    """Append a feed entry to the Excel file."""
    with file_lock:
        try:
            wb = load_workbook(get_excel_file())
            ws = wb.active

            # Parse timestamp
            if isinstance(feed_data.get("timestamp"), str):
                timestamp = parse_iso_timestamp(feed_data["timestamp"])
                # Always convert to local system time
                timestamp = timestamp.astimezone(None)
            else:
                timestamp = datetime.now().astimezone(None)

            # Format type
            feed_type_str = format_feed_type(
                feed_data.get("type"),
                feed_data.get("side")
            )

            # Build row
            row = [
                timestamp.strftime("%Y-%m-%d"),  # Date
                timestamp.strftime("%I:%M %p"),  # Time
                feed_type_str,  # Type
                feed_data.get("amount_ml"),  # Amount
                feed_data.get("duration_min"),  # Duration
                feed_data.get("notes", ""),  # Notes
                feed_data.get("logged_by", ""),  # Logged By
                timestamp.isoformat()  # Timestamp
            ]

            ws.append(row)
            wb.save(get_excel_file())

            # Return row number (excluding header)
            return ws.max_row - 1

        except Exception as e:
            print(f"Error writing to Excel: {e}")
            raise


def get_feeds_from_excel(date_filter=None, min_date=None):
    """Read feeds from Excel, optionally filtered by specific date or date range."""
    with file_lock:
        try:
            wb = load_workbook(get_excel_file())
            ws = wb.active

            feeds = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 8:
                    continue

                date_str, time_str, type_str, amount, duration, notes, logged_by, timestamp_str = row[:8]

                # Filter by specific date if requested
                if date_filter and date_str != date_filter:
                    continue

                # Filter by min_date if requested (for history range)
                if min_date and date_str < min_date:
                    continue

                feeds.append({
                    "id": idx - 1,  # Row number minus header
                    "date": date_str,
                    "time": time_str,
                    "type": type_str,
                    "amount_ml": amount,
                    "duration_min": duration,
                    "notes": notes or "",
                    "logged_by": logged_by or "",
                    "timestamp": timestamp_str
                })

            return feeds

        except Exception as e:
            print(f"Error reading Excel: {e}")
            return []


def delete_feed_from_excel(feed_id):
    """Delete a feed entry by row number."""
    with file_lock:
        try:
            wb = load_workbook(get_excel_file())
            ws = wb.active

            # feed_id is row number minus header, so actual row is feed_id + 1
            row_num = feed_id + 1

            if row_num < 2 or row_num > ws.max_row:
                return False

            ws.delete_rows(row_num)
            wb.save(get_excel_file())
            return True

        except Exception as e:
            print(f"Error deleting from Excel: {e}")
            return False


def update_feed_in_excel(feed_id, feed_data):
    """Update a feed entry by row number."""
    with file_lock:
        try:
            wb = load_workbook(get_excel_file())
            ws = wb.active

            row_num = feed_id + 1

            if row_num < 2 or row_num > ws.max_row:
                return False

            # Parse timestamp
            if isinstance(feed_data.get("timestamp"), str):
                timestamp = parse_iso_timestamp(feed_data["timestamp"])
                # Always convert to local system time
                timestamp = timestamp.astimezone(None)
            else:
                # Preserve existing timestamp if not provided
                existing_ts_str = ws[f"H{row_num}"].value
                if existing_ts_str:
                    timestamp = parse_iso_timestamp(existing_ts_str)
                else:
                    timestamp = datetime.now().astimezone(None)

            # Format type
            feed_type_str = format_feed_type(
                feed_data.get("type"),
                feed_data.get("side")
            )

            # Update row
            ws[f"A{row_num}"] = timestamp.strftime("%Y-%m-%d")
            ws[f"B{row_num}"] = timestamp.strftime("%I:%M %p")
            ws[f"C{row_num}"] = feed_type_str
            ws[f"D{row_num}"] = feed_data.get("amount_ml")
            ws[f"E{row_num}"] = feed_data.get("duration_min")
            ws[f"F{row_num}"] = feed_data.get("notes", "")
            ws[f"G{row_num}"] = feed_data.get("logged_by", "")
            ws[f"H{row_num}"] = timestamp.isoformat()

            wb.save(get_excel_file())
            return True

        except Exception as e:
            print(f"Error updating Excel: {e}")
            return False


@app.route("/")
def index():
    """Serve the main UI."""
    return render_template("index.html")


@app.route("/api/feeds", methods=["GET"])
def get_feeds():
    """Get feed entries for a specific date (defaults to today)."""
    # Check for limit_days parameter (history view)
    limit_days = request.args.get("limit_days", type=int)
    date_filter = request.args.get("date")

    if limit_days:
        # Calculate start date (Today - limit_days)
        # Note: limit_days=1 means today + yesterday (last 24h extended to specific days)
        # Actually logic: if limit=7, we want today + 6 previous days = 7 days total?
        # Or just previous N days.
        # User asked for "7 days". Let's do Today + 6 past days.
        start_date = (datetime.now() - timedelta(days=limit_days)).strftime("%Y-%m-%d")
        feeds = get_feeds_from_excel(min_date=start_date)
    elif not date_filter:
        # Default to today
        date_filter = datetime.now().strftime("%Y-%m-%d")
        feeds = get_feeds_from_excel(date_filter)
    else:
        # Specific date requested
        feeds = get_feeds_from_excel(date_filter)

    # Sort by timestamp descending (most recent first)
    feeds.sort(key=lambda x: x["timestamp"], reverse=True)

    # Calculate stats (exclude Vitamin D from feed/diaper stats)
    last_feed_minutes_ago = None
    last_feed_summary = None
    last_diaper_minutes_ago = None
    last_diaper_summary = None
    total_ml_today = 0
    total_feeds_today = 0

    if feeds:
        # Find most recent actual feed (not Vitamin D)
        last_feed = None
        for feed in feeds:
            if "Vitamin D" not in feed["type"]:
                last_feed = feed
                break

        if last_feed:
            last_timestamp = parse_iso_timestamp(last_feed["timestamp"])
            # Remove timezone info if present to avoid comparison errors
            if last_timestamp.tzinfo is not None:
                last_timestamp = last_timestamp.replace(tzinfo=None)
            last_feed_minutes_ago = int((datetime.now() - last_timestamp).total_seconds() / 60)

            # Build summary
            amount_str = f"{last_feed['amount_ml']} ml" if last_feed["amount_ml"] else ""
            duration_str = f"{last_feed['duration_min']} min" if last_feed["duration_min"] else ""
            detail_str = " — ".join(filter(None, [amount_str, duration_str]))

            last_feed_summary = f"{last_feed['type']}"
            if detail_str:
                last_feed_summary += f" — {detail_str}"
            last_feed_summary += f" at {last_feed['time']}"

        # Calculate last diaper change
        for feed in feeds:
            if "Diaper" in feed["type"]:
                last_diaper_timestamp = parse_iso_timestamp(feed["timestamp"])
                if last_diaper_timestamp.tzinfo is not None:
                    last_diaper_timestamp = last_diaper_timestamp.replace(tzinfo=None)
                last_diaper_minutes_ago = int((datetime.now() - last_diaper_timestamp).total_seconds() / 60)
                last_diaper_summary = f"{feed['type']} at {feed['time']}"
                break

        # Calculate total ml and feed count (exclude Vitamin D)
        for feed in feeds:
            if "Vitamin D" in feed["type"]:
                continue
            total_feeds_today += 1
            if feed["amount_ml"]:
                total_ml_today += feed["amount_ml"]

    return jsonify({
        "feeds": feeds,
        "last_feed_minutes_ago": last_feed_minutes_ago,
        "last_feed_summary": last_feed_summary,
        "last_diaper_minutes_ago": last_diaper_minutes_ago,
        "last_diaper_summary": last_diaper_summary,
        "total_ml_today": round(total_ml_today, 1),
        "total_feeds_today": total_feeds_today
    })


@app.route("/api/feeds", methods=["POST"])
def create_feed():
    """Log a new feed entry."""
    feed_data = request.json

    try:
        feed_id = add_feed_to_excel(feed_data)

        # Return the created entry with 201 status
        return jsonify({
            "success": True,
            "id": feed_id,
            "message": "Feed logged successfully"
        }), 201
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/feeds/<int:feed_id>", methods=["DELETE"])
def delete_feed(feed_id):
    """Delete a feed entry."""
    success = delete_feed_from_excel(feed_id)

    if success:
        return jsonify({"success": True, "message": "Feed deleted"})
    else:
        return jsonify({"success": False, "error": "Feed not found"}), 404


@app.route("/api/feeds/<int:feed_id>", methods=["PUT"])
def update_feed(feed_id):
    """Update a feed entry."""
    feed_data = request.json
    success = update_feed_in_excel(feed_id, feed_data)

    if success:
        return jsonify({"success": True, "message": "Feed updated"})
    else:
        return jsonify({"success": False, "error": "Feed not found"}), 404


@app.route("/api/vitamin-status", methods=["GET"])
def get_vitamin_status():
    """Check if Vitamin D has been given today. Also lazily auto-logs missed doses."""
    today = datetime.now().strftime("%Y-%m-%d")
    today_feeds = get_feeds_from_excel(today)

    # Check today's vitamin status
    vitamin_feed = None
    for feed in today_feeds:
        if "Vitamin D" in feed["type"]:
            vitamin_feed = feed
            break

    # Lazy missed-dose check: did yesterday have a vitamin entry?
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    yesterday_feeds = get_feeds_from_excel(yesterday)
    has_yesterday_vitamin = any("Vitamin D" in f["type"] for f in yesterday_feeds)

    if not has_yesterday_vitamin and yesterday_feeds:
        # Yesterday had feeds but no vitamin — auto-log missed dose
        yesterday_end = datetime.strptime(yesterday + " 23:59:00", "%Y-%m-%d %H:%M:%S")
        missed_data = {
            "type": "vitamin_d",
            "side": None,
            "amount_ml": None,
            "duration_min": None,
            "notes": "No",
            "logged_by": "Auto",
            "timestamp": yesterday_end.isoformat()
        }
        add_feed_to_excel(missed_data)

    if vitamin_feed:
        return jsonify({
            "given_today": True,
            "vitamin_feed_id": vitamin_feed["id"],
            "time_given": vitamin_feed["time"]
        })
    else:
        return jsonify({
            "given_today": False,
            "vitamin_feed_id": None,
            "time_given": None
        })


@app.route("/api/vitamin", methods=["POST"])
def log_vitamin():
    """Log Vitamin D administration."""
    data = request.json or {}

    feed_data = {
        "type": "vitamin_d",
        "side": None,
        "amount_ml": None,
        "duration_min": None,
        "notes": "Yes",
        "logged_by": data.get("logged_by", ""),
        "timestamp": datetime.now().isoformat()
    }

    try:
        feed_id = add_feed_to_excel(feed_data)
        return jsonify({
            "success": True,
            "id": feed_id,
            "message": "Vitamin D logged"
        }), 201
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/stats", methods=["GET"])
def get_stats():
    """Get summary statistics."""
    today = datetime.now().strftime("%Y-%m-%d")
    feeds = get_feeds_from_excel(today)

    total_ml = 0
    total_feeds = 0
    nursing_sessions = 0
    pump_ml = 0
    diaper_changes = 0

    timestamps = []

    for feed in feeds:
        # Skip Vitamin D entries from feed stats
        if "Vitamin D" in feed["type"]:
            continue

        if "Bottle" in feed["type"]:
            total_feeds += 1
            if feed["amount_ml"]:
                total_ml += feed["amount_ml"]

        if "Nurse" in feed["type"]:
            nursing_sessions += 1

        if "Pump" in feed["type"] and feed["amount_ml"]:
            pump_ml += feed["amount_ml"]

        if "Diaper" in feed["type"]:
            diaper_changes += 1

        if feed["timestamp"]:
            dt = parse_iso_timestamp(feed["timestamp"])
            if dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
            timestamps.append(dt)

    # Calculate average interval
    avg_interval = None
    if len(timestamps) > 1:
        timestamps.sort()
        intervals = []
        for i in range(1, len(timestamps)):
            interval = (timestamps[i] - timestamps[i-1]).total_seconds() / 60
            intervals.append(interval)
        avg_interval = int(sum(intervals) / len(intervals))

    return jsonify({
        "today": {
            "total_ml": round(total_ml, 1),
            "total_feeds": total_feeds,
            "total_nursing_sessions": nursing_sessions,
            "total_pump_ml": round(pump_ml, 1),
            "avg_feed_interval_min": avg_interval,
            "total_diaper_changes": diaper_changes
        }
    })


if __name__ == "__main__":
    # Initialize Excel file
    init_excel_file()

    # Get local IP
    local_ip = get_local_ip()

    # Port configuration
    PORT = 8080  # Using 8080 to avoid firewall/AirPlay conflicts on port 5000

    print("\n" + "="*60)
    print("🍼 Baby Feed Tracker")
    print("="*60)
    print(f"\n📱 Open on your phone:")
    print(f"   http://{local_ip}:{PORT}")
    print(f"\n💻 Or on this computer:")
    print(f"   http://localhost:{PORT}")
    print(f"\n📊 Data saved to: {os.path.abspath(get_excel_file())}")
    print("\nPress Ctrl+C to stop\n")
    print("="*60 + "\n")

    # Run the app
    app.run(host="0.0.0.0", port=PORT, debug=False)
